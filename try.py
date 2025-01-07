import os   #Os function for file handling
import re   #Regular expression for pattern matching
from openpyxl import Workbook, load_workbook   #To create and manipulate Excel file
import imaplib  #Connect to email server to retrieve emails
from playwright.sync_api import sync_playwright #For browser automation to fill google form
from skimage import io  #Image I/O operation
from skimage.color import rgb2gray  #Convert to grayscale
from skimage.restoration import denoise_bilateral   #Denoising image
from skimage.filters import threshold_local    #Adaptive threshold
from paddleocr import PaddleOCR    #To extract text from image
import pymupdf  #To save pdf images
import email #To parse and manage email content
import numpy as np  #Numerical operations
from email.message import EmailMessage  #Handle email message for sending
import smtplib  #Connect to email server to send emails
import pandas as pd #Excel reading

# Preprocess image for better OCR results using skimage
def preprocess_image(image_path):
    img = io.imread(image_path) #Load image from the given path
    img = rgb2gray(img) #Converts image to grayscale
    denoise = denoise_bilateral(img, sigma_color = 0.001, sigma_spatial = 5) #Denoise the image using Bilateral Denoising 
    thresh = threshold_local(denoise)   #Adaptive threshold to create binary image
    preprocessed_image_path = "D:\\preprocessed_image.png"  #Path to save processed image
    io.imsave(preprocessed_image_path, (thresh * 255).astype(np.uint8)) #Saves processed image
    return preprocessed_image_path  #Returns the processed image path

#Extract text from image using PaddleOCR
def extract_text_from_image(preprocessed_image_path):
    try:
        ocr = PaddleOCR(use_angle_cls=True, lang='en')  # Initialize PaddleOCR with English language
        results = ocr.ocr(preprocessed_image_path, cls=True)    # Perform OCR and angle correction
        extracted_text = "\n".join([line[1][0] for line in results[0]]) #Extract text
        return extracted_text   #Returns the extracted text
        
    except Exception as e:
        print(f"Error extracting text from {preprocessed_image_path}: {e}") #Handles OCR errors
        return ""   #Returns empty string in case of an error

#Extract the required information from the extracted text usig regex expression
def extract_additional_data(extracted_text):
    date_pattern = r"(\d{1,2}[-/.\s]\d{1,2}[-/.\s]\d{2,4}|\d{4}[-/.\s]\d{1,2}[-/.\s]\d{2}|\d{1,2}[-/.\s](Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/.\s]\d{4}|Date\s*?(\d{2}[-/]\d{2}[-/]\d{2})|\b\d{1,2}[-/]\d{2,4})\b"  # Match dates
    amount_pattern = r"(?i)(total|amount|grand total|TL|cash|balance\s?due|amountrs.?\)?):?\s*\$?(\d+(?:\.\d{2})?)\b"  # Match amounts
    bill_no_pattern = r"(?i)\b(bill\s?no|invo?0?ice\s?no?0?|Order\s?no|Check|Order|Inv.No?0?.?|Receipt\s?no.?|ref\s?no.?):?\s*(\w+)\b" # Match bill numbers 
    vendor_pattern = r'[A-Z][a-zA-Z0-9&\s]+(?:restaurant|hotel|petroleum|cafe)?\b'  #Match Vendor names
    date = re.search(date_pattern, extracted_text, re.IGNORECASE) #Search for date
    amount = re.search(amount_pattern, extracted_text) #Search for amount
    bill_no = re.search(bill_no_pattern, extracted_text)    #Search for bill number
    vendor = re.search(vendor_pattern, extracted_text)  #Search for vendor name
    # Store results in a dictionary
    details = {
        "date": date.group(0) if date else "Not found",
        "amount": amount.group(2) if amount else "Not found",
        "bill_number": bill_no.group(2) if bill_no else "Not found",
        "vendor" : vendor.group(0) if vendor else "Not found",
    }
    return details  #Returns the extracted data in a dictionary

#Extract text from pdf using pymupdf
def extract_text_from_pdf(pdf_path):
    retrieved_images = []   #List to store extracted images
    pdfimg = pymupdf.open(pdf_path) #Open pdf 
    for page_num in range(len(pdfimg)): #Loops through each page
        page = pdfimg[page_num] #Access current page
        pix = page.get_pixmap() #Converts page to image
        saveto = "D:\\Bills\\pdf_images\\"  #Directory to save images
        os.makedirs(saveto, exist_ok = True)    #Creates directory if it does not exist 
        file = f"{saveto}/{page_num+1}.jpg" #Filename for the image
        pix.save(file)  #Save image
        retrieved_images.append(file)   #Appends image to list
        for image in retrieved_images:  #Loops through the list
            pre = preprocess_image(image) #Preprocess the image in the list
            text = extract_text_from_image(pre) #Extracts texts fron image
            return text #Returns the extracted text

# Handle email attachments
def handle_attachment(part):
    filename = part.get_filename()  #Get the attachment filename
    attachment_data = {key: "Not found" for key in ["attachment","date", "bill_number", "amount", "vendor"]}  #Creates dictionary with keys
    if filename:
        print(f"Processing attachment: {filename}") 
        file_extension = os.path.splitext(filename)[1].lower()  #Extract file extension
        download = os.path.join("D:\\Bills\\",filename) #Saves file at given path
        os.makedirs("D:\\Bills\\", exist_ok = True) #Creates path if does not exists
        with open(download, 'wb') as f:             #Opens the file in write mode
            f.write(part.get_payload(decode=True))

        try:
            if file_extension == ".pdf":            #If pdf, calls extract_text_from_pdf function and extracts required data
                text = extract_text_from_pdf(download)
                print(f"Extracted Text:\n{text}")
                attachment_data = extract_additional_data(text) 
            elif file_extension in [".png", ".jpg", ".jpeg"]:   #If image, preprocesses the image and extracts text from it and extracts required data
                pre = preprocess_image(download)
                text = extract_text_from_image(pre)
                print(f"Extracted Text:\n{text}")
                attachment_data = extract_additional_data(text)
        except Exception as e:
            print(f"Error processing {filename}: {e}")  #Handles error if occurred
    return attachment_data

# Connect to the email server to receive mail
def connect(email_user, email_pass):
    try:
        mail = imaplib.IMAP4_SSL('imap.gmail.com')  #Connects to Gmail IMAP server
        mail.login(email_user, email_pass)  #Login with credentials
        mail.select('inbox')    #Selects the inbox
        print("Connected to email server.") #Prints message on connection
        return mail #Returns mail connection
    except Exception as e:
        print(f"Failed to connect: {e}")    #If error occurred during establishment, prints failed message
        return None 

#Connect to the email server to send mail using smtplib
def send_mail(excel_file, user, password):
    msg = EmailMessage()    #To manipulate email message
    data = pd.read_excel(excel_file)    #Reads the excel file using pandas
    body = "Receipt Details : \n\n" #Assigns body variable
    for i, row in data.iterrows():  #Loops through each row in excel file
        detail = f'Receipt {i+1}\n'
        for col, value in row.items():  #Loops through each item in excel file
            detail += f'{col} : {value}\n'  #Adds into detail the column name and its value
        detail += '\n'  #For new line
        body += detail  #Adds detail to the body
    msg.set_content(body)   #Set content with body
    msg['subject'] = 'Receipt Details'  #Set subject
    msg['to'] = "yuktigadhiya1492@gmail.com"  #Send mail to ID
    msg['from'] = user  #Send mail from ID
    #password = "tmlr kryx mjnd rtzy"
    server = smtplib.SMTP('smtp.gmail.com', 587)    #Connect to smtp
    server.starttls()   
    server.login(user, password)    #Login using given credentials
    server.send_message(msg)    
    server.quit()   #Quit the server
    print("Email sent Successfully from Excel file")    #Prints Mail sent successfully on complete execution

# Fetch emails from the inbox
def fetch_emails(mail):
    result, data = mail.search(None, 'ALL') #Search for all email in the mail box
    data_list = []  #Assign empty list  
    for num in data[0].split(): #Loops through the mail
        result, email_data = mail.fetch(num, '(RFC822)')    #Retrieves full content of an email
        raw_email = email_data[0][1].decode('utf-8')    #Decode the raw email content
        email_message = email.message_from_string(raw_email)    #Parses the string into email_message
        data_list.append(extract_email_data(email_message)) #Appends the list with extracted data from email_message
    return data_list    #Returns data_list

# Extract data from email
def extract_email_data(email_message):
    attachment = "No"
    attachment_data = {key: "Not found" for key in ["attachment", "date", "bill_number", "amount", "vendor"]}   
    for part in email_message.walk():   #Loop through each content of body
        if part.get_content_disposition() == "attachment":  #If attachment found
            attachment = "Yes"
            attachment_data = handle_attachment(part)   #Calls the handle_attachment function and saves the extracted data in the variable
            break
    return [attachment, attachment_data['date'], attachment_data['bill_number'], attachment_data['amount'], attachment_data['vendor']]  #Returns the value of the dictionary

# Save extracted data to Excel using openpyxl
def save_to_excel(data_list, filename):
    wb = Workbook() #Creates the workbook
    ws = wb.active  #Gets the currently active workbook
    ws.append(["Attachment", "Date", "Bill Number", "Amount", "Vendor"])    #Adds column
    for data in data_list:  #Loops through each data in data_list
        ws.append(data) #Appends the data in the workbook
    wb.save(filename)   #Saves the workbook
    workbook = load_workbook(filename)      #Loads the workbook in the terminal
    sheet = workbook.active #Gets the workbook
    print("\nExcel File Content:")
    for row in sheet.iter_rows(values_only=True):   #Iterates through each row
        print(row)  #Prints the content of the row
    return filename #Returns the filename

# Fill Google Form using playwright
def fill_google_form(excel_file, form_url):
    try:
        data = pd.read_excel(excel_file)    #Reads the excel file using pandas
        with sync_playwright() as p:    #Assigns sync_playwright function to p
            browser = p.chromium.launch(headless=False,executable_path="C:/Program Files/Google/Chrome/Application/chrome.exe")  #Launches the browser
            page = browser.new_page()   #Launches the browser in a new page
            page.goto(form_url) #Directs to the given form_url
            for i, row in data.iterrows():  #Iterates over rows present in excel sheet
                print(f"Filling form for row {i + 1}")  
                page.fill("input[aria-labelledby='i1 i4']", row['Attachment'])  #Fills the value at given position 
                page.fill("textarea[aria-labelledby='i6 i9']", row['Date']) 
                page.fill("input[aria-labelledby='i11 i14']", row['Bill Number'])  
                page.fill("textarea[aria-labelledby='i16 i19']", str(row['Amount']))  
                page.fill("input[aria-labelledby='i21 i24']", row['Vendor'])    
                page.click("span:has-text('Submit')")
                print(f"Form filled {i+1}")
                page.goto(form_url) #Redirects to the form_url until the form for all the rows are filled 
            browser.close() #Closes the browser
    except Exception as e:
            print(f"Error filling form for row {i}: {e}")
#To show windows notification
def windows_notification(title, message):
    command = f'PowerShell -Command "New-BurntToastNotification -Text \\"{title}\\", \\"{message}\\""'  #Shows notification using powershell's burnt toast notification
    os.system(command)  #Shows notification

# Main function
if __name__ == "__main__":
    email_user = "yukti2792@gmail.com"
    email_pass = "tmlr kryx mjnd rtzy"  
    mail = connect(email_user, email_pass)  #Connects to the server using given credentials

    if mail:
        data_list = fetch_emails(mail)  #Calls fetch_emails function and assigns the output to the data_list
        excel_file = "extracted_data.xlsx"
        save = save_to_excel(data_list, excel_file) #Saves the data_list at given file
        form_url ="https://docs.google.com/forms/d/e/1FAIpQLSeBljdidu_NCEfrMXoQqznf7aVN2m5A21k5GDfp423Lq4xZ3A/viewform?usp=header"
        form = fill_google_form(excel_file, form_url)   #Fills the google form using the given form_url and excel file
        send_mail(excel_file, email_user, email_pass)   #Sends the mail 
        title = "Process Done!!"
        message = "Data Extracted and Stored Successfully!!"
        windows_notification(title, message)    #Windows notification pops up after completion
        mail.logout()   #Disconnects/Logout from the server


